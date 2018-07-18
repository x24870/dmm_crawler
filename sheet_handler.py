import openpyxl, os
import dmm_crawler

CELL_SIZE = 50

def create_header(wb):
    t_font = openpyxl.styles.Font(name='Times New Roman', bold=True)

    #Set header
    sheet = wb.active
    sheet.cell(1, 1).value = 'Title'
    sheet.cell(1, 1).font = t_font
    sheet.cell(1, 2).value = 'Cover'
    sheet.cell(1, 2).font = t_font
    sheet.cell(1, 3).value = 'Sample'
    sheet.cell(1, 3).font = t_font
    
    #Adjust cell size
    sheet.column_dimensions['B'].width = 10

def insert_img(wb, img_name, row_idx, col_idx):
    sheet = wb.active
    img = openpyxl.drawing.image.Image(img_name)
    sheet.add_image(img, openpyxl.utils.get_column_letter(col_idx)+str(row_idx))
    img.width = CELL_SIZE
    img.height = CELL_SIZE
    sheet.row_dimensions[row_idx].height = 50

def insert_title(wb, work_name, idx):
    sheet = wb.active
    sheet.cell(idx+2, 1).value = work_name

def insert_cover_img(wb, work_name, idx):
    img_dir = os.path.join(os.path.join(dmm_crawler.PIC_DIR, work_name), 'cover')
    img_name = os.path.join(img_dir, os.listdir(img_dir)[0])
    insert_img(wb, img_name, idx+2, 2)

def insert_sample_img(wb, work_name, idx):
    img_dir = os.path.join(os.path.join(dmm_crawler.PIC_DIR, work_name), 'sample')
    img_names = os.listdir(img_dir)

    for img_idx, img_name in enumerate(img_names):
        name = os.path.join(img_dir, img_name)
        insert_img(wb, name, idx+2, img_idx+3)
    

def test():
    wb = openpyxl.Workbook()

    create_header(wb)

    works = os.listdir(dmm_crawler.PIC_DIR)
    for idx, work in enumerate(works):
        insert_title(wb, work, idx)
        insert_cover_img(wb, work, idx)
        insert_sample_img(wb, work, idx)

    #insert_title(wb, 'hamenets102', 0)
    #insert_cover_img(wb, 'hamenets102', 0)
    #insert_sample_img(wb, 'hamenets102', 0)

    wb.save('sheet.xlsx')

if __name__ == '__main__':
    test()